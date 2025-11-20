#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Restructure Investing.com macro calendar export into a flat CSV.

This version preserves the visual percent formatting from Excel
(e.g. 3.02%) instead of writing 0.0302.
"""

from __future__ import annotations

from pathlib import Path
from typing import Final, Optional, Tuple, List, Dict, Any
from datetime import datetime
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

# =========================
# CONFIG
# =========================

INPUT_XLSX: Final[Path] = Path(r"C:\Users\SM\OneDrive\Documents\Binus\Thesis\skripsi_2602214545\data\raw\investing_com_id_econ_cal_unstructured.xlsx")
OUTPUT_CSV_ALL: Final[Path] = Path(r"C:\Users\SM\OneDrive\Documents\Binus\Thesis\skripsi_2602214545\data\raw\investing_com_id_econ_cal_structured.csv")


# Columns we expect in the Investing.com export
EXPECTED_COLS: Final[List[str]] = [
    "Time",
    "Currency",
    "Impact",
    "Event",
    "Actual",
    "Forecast",
    "Previous",
]


def looks_like_date_cell(val: Optional[str]) -> bool:
    """
    Detect if val is like 'Thursday, January 2, 2020'.
    We'll try multiple datetime formats.
    """
    if val is None:
        return False
    s: str = str(val).strip()
    if not s:
        return False

    fmts: List[str] = [
        "%A, %B %d, %Y",   # Thursday, January 2, 2020
        "%A, %b %d, %Y",   # Thu, Jan 2, 2020
        "%A %B %d, %Y",    # Thursday January 2, 2020
        "%B %d, %Y",       # January 2, 2020
    ]

    for fmt in fmts:
        try:
            datetime.strptime(s, fmt)
            return True
        except ValueError:
            continue
    return False


def parse_date_string(val: str) -> str:
    """
    Convert e.g. 'Thursday, January 2, 2020' -> '2020-01-02'
    """
    s: str = str(val).strip()
    fmts: List[str] = [
        "%A, %B %d, %Y",
        "%A, %b %d, %Y",
        "%A %B %d, %Y",
        "%B %d, %Y",
    ]
    for fmt in fmts:
        try:
            dt: datetime = datetime.strptime(s, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue

    raise ValueError(f"Unrecognized date format: {val!r}")


def format_cell_value(cell: Cell) -> Optional[str]:
    """
    Return the *display-style* value we want for CSV:
    - If the cell is numeric AND the number format looks like percent,
      convert 0.0302 -> "3.02%".
    - Else just return string(cell.value), except empty -> None.
    """

    v: Any = cell.value

    # Handle blanks
    if v is None:
        return None
    if isinstance(v, str):
        s_clean: str = v.strip()
        return s_clean if s_clean != "" else None

    # Handle numeric cells
    if isinstance(v, (int, float)):
        number_format: str = str(cell.number_format).lower()

        # Heuristic: if Excel formatted it as a percentage, rebuild it
        if "0%" in number_format or "%" in number_format:
            # v is like 0.0302, so multiply by 100
            pct_val: float = float(v) * 100.0
            # Try to keep 2 decimal places by default
            return f"{pct_val:.2f}%"

        # Not a percent-format cell; just stringify it
        # but avoid scientific notation surprises
        return f"{v}"

    # Fallback for other types (datetime, etc.) -> string
    return str(v)


def read_investing_xlsx(path: Path) -> pd.DataFrame:
    """
    We CANNOT just do pd.read_excel if we want to preserve percent signs,
    because Excel stores 3.02% as numeric 0.0302 + a percent cell format.

    So:
    - openpyxl load
    - walk rows manually
    - build a pandas.DataFrame
    """

    wb = load_workbook(filename=path, data_only=True)
    ws = wb.active  # assume first sheet

    rows_as_dicts: List[Dict[str, Optional[str]]] = []

    # First, infer header names from the first row that matches EXPECTED_COLS subset.
    header_map: Dict[int, str] = {}
    header_found: bool = False

    for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
        # Build a candidate header row (stringified)
        candidate_headers: List[str] = [ (str(c.value).strip() if c.value is not None else "") for c in row ]
        # Simple heuristic: if it contains "Time" and "Event", assume it's the header row
        if ("Time" in candidate_headers) and ("Event" in candidate_headers) and (not header_found):
            header_found = True
            for col_idx, name in enumerate(candidate_headers):
                if name in EXPECTED_COLS:
                    header_map[col_idx] = name
            # After setting header_map, continue to next row to start reading data
            continue

        # If header not found yet, skip rows before header
        if not header_found:
            continue

        # After header is found, we interpret each subsequent row using header_map
        row_dict: Dict[str, Optional[str]] = {}
        for col_idx, cell in enumerate(row):
            if col_idx in header_map:
                col_name: str = header_map[col_idx]
                row_dict[col_name] = format_cell_value(cell)

        rows_as_dicts.append(row_dict)

    # Now build DataFrame
    df: pd.DataFrame = pd.DataFrame(rows_as_dicts, columns=EXPECTED_COLS)
    return df


def build_flat_dataframe(raw_df: pd.DataFrame) -> pd.DataFrame:
    """
    Flatten the "date header row + multiple event rows" pattern
    into a normal table with 'Date' column.
    """
    records: List[Dict[str, Optional[str]]] = []
    current_date: Optional[str] = None

    # We assume the date header text shows up in the first column of raw_df,
    # which should correspond to "Time" after parsing.
    first_col_name: str = "Time"

    for _, row in raw_df.iterrows():
        first_val: Optional[str] = row.get(first_col_name, None)

        # Check if this row is actually a date header
        if isinstance(first_val, str) and looks_like_date_cell(first_val):
            current_date = parse_date_string(first_val)
            continue

        if current_date is None:
            # skip any lines before seeing first valid date header
            continue

        rec: Dict[str, Optional[str]] = {
            "Date": current_date,
            "Time": row.get("Time", None),
            "Currency": row.get("Currency", None),
            "Impact": row.get("Impact", None),
            "Event": row.get("Event", None),
            "Actual": row.get("Actual", None),
            "Forecast": row.get("Forecast", None),
            "Previous": row.get("Previous", None),
        }

        # Detect fully empty rows (sometimes appear in export)
        is_empty: bool = all(
            (rec[k] is None or (isinstance(rec[k], str) and rec[k].strip() == ""))
            for k in ["Time", "Currency", "Impact", "Event", "Actual", "Forecast", "Previous"]
        )
        if is_empty:
            continue

        records.append(rec)

    flat_df: pd.DataFrame = pd.DataFrame.from_records(records)
    return flat_df


def normalize_time_to_sortable(t: Optional[str]) -> Tuple[int, int, int]:
    """
    Turn Time into a sorting key (HH, MM, SS).
    'All Day' -> (0,0,0) so it floats to top of that date.
    Unknown/blank -> put it last.
    """
    if t is None or t.strip() == "":
        return (23, 59, 59)

    s: str = t.strip()

    if re.match(r"(?i)^all\s*day$", s):
        return (0, 0, 0)

    try:
        hh_str, mm_str = s.split(":")
        hh: int = int(hh_str)
        mm: int = int(mm_str)
        return (hh, mm, 0)
    except Exception:
        return (23, 59, 59)


def main() -> None:
    # 1. Load workbook + preserve display formatting (especially %)
    raw_df: pd.DataFrame = read_investing_xlsx(INPUT_XLSX)

    # 2. Flatten date-grouped rows into (Date, Time, Currency, ...)
    flat_df: pd.DataFrame = build_flat_dataframe(raw_df)

    # 3. Sort rows by Date then Time
    sort_keys = flat_df["Time"].apply(normalize_time_to_sortable)
    flat_df["_sort_tuple"] = sort_keys
    flat_df = flat_df.sort_values(by=["Date", "_sort_tuple"], kind="stable").drop(columns=["_sort_tuple"])

    # 4. Write to CSV (everything stays as string, including percents)
    OUTPUT_CSV_ALL.parent.mkdir(parents=True, exist_ok=True)
    flat_df.to_csv(OUTPUT_CSV_ALL, index=False)

    print(f"Saved full calendar (no filtering): {len(flat_df)} rows -> {OUTPUT_CSV_ALL}")


if __name__ == "__main__":
    main()
